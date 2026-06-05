"""
VOXScript - Gemini API Formatter
"""

import os
import re
import time
from dataclasses import dataclass
from DAO.translator import TranslationResult
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT_FORMATS = {
    "txt": "plain text (translation only)",
    "txt_bilingual": "text with timecode (original + translation)",
    "srt": "SRT subtitle (translation only)",
    "srt_bilingual": "SRT subtitle bilingual (original + translation)",
    "excel": "Excel (timecode / original / translation)",
    "all": "save all formats",
}


@dataclass
class FormatResult:
    saved_files: list[str]
    claude_summary: str | None = None
    auto_title: str | None = None


DEFAULT_EXPORT_DIR = Path.home() / "Downloads" / "VOXScript" / "output"


def format_and_save(
    translation_result: TranslationResult,
    output_name: str,
    formats: list[str],
    export_dir: Path | None = None,
    use_claude_summary: bool = True,
    progress_callback=None,
) -> FormatResult:
    base_dir = Path(export_dir) if export_dir else DEFAULT_EXPORT_DIR

    # 임시 폴더
    temp_dir = base_dir / "_temp"
    temp_dir.mkdir(parents=True, exist_ok=True)

    temp_name = f"vox_{int(time.time())}"
    temp_files: dict[str, Path] = {}

    if "all" in formats:
        formats = ["txt", "txt_bilingual", "srt", "srt_bilingual", "excel"]

    if "txt" in formats:
        path = temp_dir / f"{temp_name}_번역.txt"
        path.write_text(translation_result.to_plain_text(), encoding="utf-8")
        temp_files["txt"] = path

    if "txt_bilingual" in formats:
        path = temp_dir / f"{temp_name}_원문번역.txt"
        path.write_text(translation_result.to_bilingual_text(), encoding="utf-8")
        temp_files["txt_bilingual"] = path

    if "srt" in formats:
        path = temp_dir / f"{temp_name}_번역.srt"
        path.write_text(translation_result.to_translated_only_srt(), encoding="utf-8")
        temp_files["srt"] = path

    if "srt_bilingual" in formats:
        path = temp_dir / f"{temp_name}_병기.srt"
        path.write_text(translation_result.to_bilingual_srt(), encoding="utf-8")
        temp_files["srt_bilingual"] = path

    if "excel" in formats:
        path = temp_dir / f"{temp_name}.xlsx"
        _save_excel(translation_result, path)
        temp_files["excel"] = path

    # Gemini 요약 + 제목 추출
    claude_summary = None
    auto_title = None

    if use_claude_summary:
        if progress_callback:
            progress_callback("Gemini summary + title extracting...", 92)
        try:
            claude_summary, auto_title = _summarize_and_get_title(
                translation_result, output_name
            )
        except Exception as e:
            print(f"[Formatter] Summary failed (skipped): {e}")

    # 폴더명은 원본 파일명 우선, Gemini 제목은 요약 파일명에만 사용
    final_name = output_name
    final_dir = base_dir / final_name
    final_dir.mkdir(parents=True, exist_ok=True)
    print(f"[Formatter] Output folder: {final_dir}")

    suffix_map = {
        "txt": f"{final_name}_번역.txt",
        "txt_bilingual": f"{final_name}_원문번역.txt",
        "srt": f"{final_name}_번역.srt",
        "srt_bilingual": f"{final_name}_병기.srt",
        "excel": f"{final_name}.xlsx",
    }

    saved = []
    for key, temp_path in temp_files.items():
        new_path = final_dir / suffix_map[key]
        temp_path.rename(new_path)
        saved.append(str(new_path))
        print(f"[Formatter] Saved: {new_path}")

    if claude_summary:
        summary_path = final_dir / f"{final_name}_요약.txt"
        summary_path.write_text(claude_summary, encoding="utf-8")
        saved.append(str(summary_path))
        print(f"[Formatter] Summary: {summary_path}")

    # 임시 폴더 정리
    try:
        temp_dir.rmdir()
    except OSError:
        pass

    return FormatResult(
        saved_files=saved, claude_summary=claude_summary, auto_title=auto_title
    )


def _sanitize_filename(name: str) -> str:
    name = re.sub(r'[\\/:*?"<>|]', "", name)
    name = name.strip().replace(" ", "_")
    return name[:50] if len(name) > 50 else name


def _extract_speaker(text: str) -> tuple[str, str]:
    """[화자] 텍스트 에서 화자와 본문 분리"""
    import re

    m = re.match(r"^\[([^\]]+)\]\s*(.*)", text.strip(), re.DOTALL)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return "", text.strip()


def _save_excel(result: TranslationResult, path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Script"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", name="Malgun Gothic")
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 한국어 소스면 Translation 컬럼 생략
    is_korean = result.source_language.lower() in ("ko", "korean")

    # 화자 컬럼이 있는지 확인
    has_speaker = any(
        _extract_speaker(seg.translated.strip())[0] for seg in result.segments
    )

    if has_speaker and not is_korean:
        headers = ["#", "Start", "End", "Speaker", "Original", "Translation"]
        col_widths = [5, 13, 13, 15, 45, 45]
    elif has_speaker and is_korean:
        headers = ["#", "Start", "End", "Speaker", "Original"]
        col_widths = [5, 13, 13, 15, 90]
    elif not has_speaker and not is_korean:
        headers = ["#", "Start", "End", "Original", "Translation"]
        col_widths = [5, 13, 13, 50, 50]
    else:  # 한국어 + 화자 없음
        headers = ["#", "Start", "End", "Original"]
        col_widths = [5, 13, 13, 100]

    # 화자별 색상 매핑
    speaker_colors = {}
    color_pool = ["EBF5FB", "E9F7EF", "FEF9E7", "F9EBEA", "F4ECF7", "FDFEFE"]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 22

    for seg in result.segments:
        row = seg.index + 2

        if has_speaker:
            speaker, orig_text = _extract_speaker(seg.original.strip())
            _, trans_text = _extract_speaker(seg.translated.strip())

            if speaker and speaker not in speaker_colors:
                color_idx = len(speaker_colors) % len(color_pool)
                speaker_colors[speaker] = color_pool[color_idx]

            row_color = speaker_colors.get(speaker, "FFFFFF")
            if is_korean:
                values = [
                    seg.index + 1,
                    seg.start_srt(),
                    seg.end_srt(),
                    speaker,
                    orig_text,
                ]
            else:
                values = [
                    seg.index + 1,
                    seg.start_srt(),
                    seg.end_srt(),
                    speaker,
                    orig_text,
                    trans_text,
                ]
        else:
            row_color = "F5F8FC" if row % 2 == 0 else "FFFFFF"
            if is_korean:
                values = [
                    seg.index + 1,
                    seg.start_srt(),
                    seg.end_srt(),
                    seg.original.strip(),
                ]
            else:
                values = [
                    seg.index + 1,
                    seg.start_srt(),
                    seg.end_srt(),
                    seg.original.strip(),
                    seg.translated.strip(),
                ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin
            is_text_col = col >= (5 if has_speaker else 4)
            cell.alignment = Alignment(vertical="top", wrap_text=is_text_col)
            cell.fill = PatternFill("solid", fgColor=row_color)

            # 화자 컬럼 강조
            if has_speaker and col == 4 and val:
                cell.font = Font(bold=True, name="Malgun Gothic")
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    wb.save(path)


def _summarize_and_get_title(result: TranslationResult, hint: str) -> tuple[str, str]:
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        raise EnvironmentError("GEMINI_API_KEY not found.")

    import google.genai as genai

    client = genai.Client(api_key=api_key)

    segs = result.segments[:200]
    full_text = "\n".join(f"[{seg.index+1}] {seg.translated.strip()}" for seg in segs)
    truncated = len(result.segments) > 200

    prompt = f"""다음은 영상의 번역 스크립트입니다.
{'(전체 내용이 길어 앞부분만 포함되었습니다)' if truncated else ''}

---
{full_text}
---

아래 형식으로 정확히 응답하세요 (TITLE 줄 필수, 첫 번째 줄):

TITLE: 영상 제목 (한국어, 20자 이내, 특수문자 제외)

1. **전체 요약** (3~5줄)
2. **주요 토픽** (bullet point, 5개 이내)
3. **핵심 발언** (인상적인 발언 3개, 원문 인용 없이 내용 요약)
4. **추가 참고사항** (번역 시 유의할 표현이나 맥락 등, 있으면)"""

    response = client.models.generate_content(model="gemini-2.5-flash", contents=prompt)
    text = response.text.strip()

    title = hint
    lines = text.splitlines()
    for line in lines:
        if line.startswith("TITLE:"):
            title = line.replace("TITLE:", "").strip()
            break

    summary = "\n".join(l for l in lines if not l.startswith("TITLE:")).strip()
    return summary, title
